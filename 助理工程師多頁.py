import requests
import bs4
import openpyxl

# 創建 Excel 工作簿
wb = openpyxl.Workbook()
sheet = wb.active

# 寫入標題欄位
sheet.cell(row=1, column=1, value='公司名稱')
sheet.cell(row=1, column=2, value='公司類別')
sheet.cell(row=1, column=3, value='地址')
sheet.cell(row=1, column=4, value='薪資')
sheet.cell(row=1, column=5, value='網址')

# 要爬取的頁數範圍
start_page = 1import requests
import bs4
import openpyxl

# 創建 Excel 工作簿
wb = openpyxl.Workbook()
sheet = wb.active

# 寫入標題欄位
sheet.cell(row=1, column=1, value='公司名稱')
sheet.cell(row=1, column=2, value='公司類別')
sheet.cell(row=1, column=3, value='地址')
sheet.cell(row=1, column=4, value='薪資')
sheet.cell(row=1, column=5, value='網址')

# 要爬取的頁數範圍
start_page = 1
end_page = 5

# 迴圈爬取多頁
for page in range(start_page, end_page + 1):
    url = f"https://www.104.com.tw/jobs/search/?ro=0&kwop=7&keyword=%E5%8A%A9%E7%90%86%E5%B7%A5%E7%A8%8B%E5%B8%AB&order=1&page={page}&mode=s&jobsource=2018indexpoc"
    res = requests.get(url)
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

    jobs = soup.find_all('article', class_='js-job-item')

    row = (page - 1) * len(jobs) + 2  # 計算寫入的起始行數

    for job in jobs:
        company = job.get('data-cust-name')
        company_type = job.get('data-indcat-desc')
        company_loc = job.find('ul', class_='job-list-intro').find('li').text
        try:
            job_pay = job.find('span', class_='b-tag--default').text
        except:
            job_pay = job.find('a', class_='b-tag--default').text
        job_url = job.find('a')['href']

        sheet.cell(row=row, column=1, value=company)
        sheet.cell(row=row, column=2, value=company_type)
        sheet.cell(row=row, column=3, value=company_loc)
        sheet.cell(row=row, column=4, value=job_pay)
        sheet.cell(row=row, column=5, value=job_url)

        row += 1

# 儲存 Excel 檔案
wb.save('job_data4.xlsx')

print("爬取完成，結果已保存到 job_data4.xlsx")
end_page = 5

# 迴圈爬取多頁
for page in range(start_page, end_page + 1):
    url = f"https://www.104.com.tw/jobs/search/?ro=0&kwop=7&keyword=%E5%8A%A9%E7%90%86%E5%B7%A5%E7%A8%8B%E5%B8%AB&order=1&page={page}&mode=s&jobsource=2018indexpoc"
    res = requests.get(url)
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

    jobs = soup.find_all('article', class_='js-job-item')

    row = (page - 1) * len(jobs) + 2  # 計算寫入的起始行數

    for job in jobs:
        company = job.get('data-cust-name')
        company_type = job.get('data-indcat-desc')
        company_loc = job.find('ul', class_='job-list-intro').find('li').text
        try:
            job_pay = job.find('span', class_='b-tag--default').text
        except:
            job_pay = job.find('a', class_='b-tag--default').text
        job_url = job.find('a')['href']

        sheet.cell(row=row, column=1, value=company)
        sheet.cell(row=row, column=2, value=company_type)
        sheet.cell(row=row, column=3, value=company_loc)
        sheet.cell(row=row, column=4, value=job_pay)
        sheet.cell(row=row, column=5, value=job_url)

        row += 1

# 儲存 Excel 檔案
wb.save('job_data4.xlsx')

print("爬取完成，結果已保存到 job_data4.xlsx")