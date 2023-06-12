import requests
import bs4
import openpyxl

# 發送 HTTP GET 請求獲取網頁內容
url = "https://www.104.com.tw/jobs/search/?ro=0&keyword=軟體助理工程師&expansionType=area%2Cspec%2Ccom%2Cjob%2Cwf%2Cwktm&order=15&asc=0&page=1&mode=s&jobsource=2018indexpoc&langFlag=0&langStatus=0&recommendJob=1&hotJob=1"
res = requests.get(url)

# 創建 BeautifulSoup 物件解析網頁內容
soup = bs4.BeautifulSoup(res.text, 'html.parser')

# 尋找所有職位元素
jobs = soup.find_all('article', class_='js-job-item')

# 創建 Excel 工作簿
wb = openpyxl.Workbook()
sheet = wb.active

# 寫入標題欄位
sheet.cell(row=1, column=1, value='公司名稱')
sheet.cell(row=1, column=2, value='公司類別')
sheet.cell(row=1, column=3, value='地址')
sheet.cell(row=1, column=4, value='薪資')
sheet.cell(row=1, column=5, value='網址')

# 寫入職位資料
row = 2
for job in jobs:
    company = job.get('data-cust-name')
    company_type = job.get('data-indcat-desc')
    company_loc = job.find('ul', class_='job-list-intro').find('li').text
    job_pay = job.find('a', class_='b-tag--default').text
    job_url = job.find('a')['href']

    sheet.cell(row=row, column=1, value=company)
    sheet.cell(row=row, column=2, value=company_type)
    sheet.cell(row=row, column=3, value=company_loc)
    sheet.cell(row=row, column=4, value=job_pay)
    sheet.cell(row=row, column=5, value=job_url)

    row += 1

# 儲存 Excel 檔案
wb.save('job_data1.xlsx')

print("爬取完成，結果已保存到 job_data1.xlsx")
